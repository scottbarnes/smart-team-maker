import sys, os, argparse, logging
from dataclasses import dataclass
from openpyxl import load_workbook
from typing import List

# Setup the workbook
workbook = load_workbook(filename="smart.xlsx")  # read_only=True breaks iter_rows.  
sheet = workbook.active


##  Column mapping (zero-indexed)
# Participants
FIELD = 24
NATIONALITY = 22
FIRST_NAME = 2 
LAST_NAME= 9
EMAIL = 1

@dataclass
class Team:
    id: int
    members: list

    def fields(self):
        result = []
        for member in self.members:
            result.append(member.field)
        return result

    def nationalities(self):
        result = []
        for member in self.members:
            result.append(member.nationality)
        return result

    def member_count(self):
        return len(self.members)


@dataclass
class Participant:
    id: int
    field: str
    nationality: str
    first_name: str
    last_name: str
    email: str


def team_creator(spreadsheet, team_size) -> List:
    """ Take a spreadsheet and a number of members per team and return a list of 
    teams with the number of members defined by team_size. """
    total_participants = get_real_max_rows(spreadsheet) + team_size  # Add another team to hold the remainder
    total_teams = total_participants // team_size  # number of teams
    result = []
    # Loop total_teams to create the team objects.
    i = 1  # start at 1
    while i < (total_teams + 1):
        result.append(Team(id=i, members=[]))
        i += 1
    return result

# Create a list of participants
def get_participants(spreadsheet) -> List:
    """ Take an openpyxl sheet and return a list of participnats, one per row. """
    result = []
    for index, row in enumerate(spreadsheet.iter_rows(min_row=1,
                                                values_only=True)):
        # Filter out the empty rows Google Docs creates.
        if row[0] == None:
            continue

        result.append(
            Participant(id=index,
                       field=row[FIELD],
                       nationality=nationality_normalizer(row[NATIONALITY]),
                       first_name=row[FIRST_NAME],
                       last_name=row[LAST_NAME],
                       email=row[EMAIL])
        )
    return result

def nationality_normalizer(nationality: str) -> str:
    """ Take a nationality string and return a normalized nationality.
        E.g. Taiwan -> Taiwanese, R.O.C. -> Taiwanese """

    nationality = str(nationality).lower()
    if 'bangl'.lower() in nationality:
        return 'Bangladeshi'
    elif 'india'.lower() in nationality:
        return 'Indian'
    elif 'indo'.lower() in nationality:
        return 'Indonesian'
    elif 'mala'.lower() in nationality:
        return 'Malaysian'
    elif 'pak'.lower() in nationality:
        return 'Pakistani'
    elif 'roc'.lower() in nationality or 'r.o.c.'.lower() in nationality or 'taiw'.lower() in nationality or 'republic of ch'.lower() in nationality:
        return 'Taiwanese'
    # China must come after because first match wins, and 'china' would match 'Republic of China'
    elif 'china'.lower() or 'chinese' in nationality:
        return 'Chinese'
    elif 'sing'.lower() in nationality:
        return 'Singaporean'
    elif 'viet'.lower() in nationality:
        return 'Vietnamese'
    elif 'usa'.lower() in nationality or 'america'.lower() in nationality:
        return 'USA'
    else:
        return f'OTHER: {nationality}' 

# Put the Participants on to teams
def inject_participant(team, participant, team_size, rnd) -> bool:
    """ For a given team and participant, return True if a participant should be injected. """

    # Don't add more than team_size members. 
    if team.member_count() >= team_size:
        return False

    # Try to make things as unique as possible.
    if rnd <= 5:
        # Fill teams with two Indian participants, owing to their enthusiasm.
        if participant.nationality == 'Indian' and (team.nationalities().count('Indian') <= 1):
            print(f'Adding Indian {participant.first_name}')
            return True
    elif rnd >= 6 and rnd <= 10:
        if participant.field in team.fields():
            return False
        elif participant.nationality in team.nationalities():
            return False
        else:
            return True
    # Now allow some duplicates
    elif rnd >= 11 and rnd <= 15:
        if participant.field == 'Engineering' and (team.fields().count('Engineering') <= 3):
            return True
        # # Add more Indians per team, owing to Indian students' overwhelming enthusiasm.
        # elif participant.nationality == 'Indian' and (team.nationalities().count('Indian') <= 3):
        #     return True
        else:
            return False
    # Now allow triplicates
    elif rnd >= 16 and rnd <= 20:
        if team.fields().count(participant.field) <= 4:
            return True
        elif team.nationalities().count(participant.nationality) <= 4:
            return True
        else:
            return False
    # Just add the first person we encounter
    elif rnd >= 21:
        return True

    return False  # Just to squelch type checking in the function definition.

    
def get_real_max_rows(spreadsheet):
    """ Take a spreadsheet, count until encountering a null row, and return that number.
    Spreadsheets from Google Docs seem to add +100 rows. Filter those out. """
    count = 0
    for row in spreadsheet.iter_rows(values_only=True):
        if row[0] != None:
            count += 1
    return count

def pretty_print_teams(teams):
    """ Take a list of teams and pretty-print them. """
    for team in teams:
        print(f'Team ID: {team.id}')
        print(f'Member count: {team.member_count()}')
        for member in team.members:
            print(f'- {member.first_name} {member.last_name}\tNationality: {member.nationality}\tField: {member.field}')


class Template:
    """
    Brief description for this class.
    """

    def __init__(self):

        # This is where you set default values and constants. Or just to declare class variables that
        # are initialized to a default value here, but will be manipulated by the functions later.

        # For example, if you want to default the work directory for this script to a tmp directory
        # somewhere, so any output or log files you generate will be easily organized and found, you
        # can set it here.

        # If you use windows, you can specify an absolute directory this way
        # self.work_dir = 'C:\\Documents and Settings\\myname' <-- replace myname with a real directory

        self.work_dir = '.'


    def get_options(self, argv):
        """
        Get command line or conf file options
        """
        p = argparse.ArgumentParser(description='')

        p.add_argument('--subdir', action='store', default=None, dest='subdir',
                       help='[None] If specified, create a subdirectory in the default work directory.')
        p.add_argument('--verbose', action='store_true', default=False, dest='verbose',
                       help='[False] If true, print verbose debug messages')

        opt = p.parse_args()

        #=======================================================================
        # Check and store opt arguments

        if opt.subdir:
            self.work_dir = os.path.join(self.work_dir, opt.subdir)
        if not os.path.exists(self.work_dir):
            os.makedirs(self.work_dir)

        self.verbose = opt.verbose
        if self.verbose:
            logging.basicConfig(level=logging.DEBUG)
        else:
            logging.basicConfig(level=logging.INFO)


    def go(self, argv):
        """
        """
        self.get_options(argv)
        self.team_size = 5

        # Now do something else.
        logging.debug("This message is only printed if self.verbose==True.")
        logging.info("This message is always printed.")
        logging.debug("self.work_dir is %s", self.work_dir)

        # Set everything up.
        participants = get_participants(sheet)
        teams = team_creator(sheet, self.team_size)
        full_teams = []
        popped = 0
        participants_popped = 0

        # Use different 'rounds' to fill the teams, taking each team in successing, and running
        # inject_participant() on it. Do this until we run out of teams (or participants).
        #
        # One thing to look out for here is that pooping the teams changes the list index,
        # So depending on how the loop is used, that could potentially break things. E.g.,
        # that is why the remaining teams, once participants run out, are appended to full_teams
        # and not .pop()ed from teams: if there were 4 teams left that weren't full, the for
        # loop would pop the first two, the last two would move two index places towards zero,
        # and the for loop would hit an empty index on teams 3 and 4.

        rnd = 0
        while True:
            # Take each team in turn, and then go through all participants trying to add one
            # until a participant is added, stop, and go to the next team. Repeat until no
            # more participants to place.
            for team in teams:
                for participant in participants:
                    if inject_participant(team, participant, self.team_size, rnd):
                        print(f'Adding {participant} to {team.id}')
                        team.members.append(participant)
                        participants.pop(participants.index(participant))
                        participants_popped += 1
                        print(f'participants remaining: {len(participants)}')
                        print(f'participants popped: {participants_popped}')

                        # Get out if successfully adding a participant and don't add more.
                        break

                if team.member_count() >= self.team_size:
                    # print(f'Appending full team: {team}')
                    full_teams.append(team)
                    print(f'Popping team: {team.id}')
                    teams.pop(teams.index(team))
                    popped += 1
                    print(f'Teams popped: {popped}')
                    print(f'Teams remaining: {len(teams)}')

            if len(participants) == 0 or len(teams) == 0:
                print(f'Popping remaining teams')
                for team in teams:
                    print(f'Popping team {team.id} with {team.member_count()} participants')
                    print(f'{team.members}')
                    print('\n')
                    full_teams.append(team)
                    # teams.pop(teams.index(team))
                    print(f'Remaining teams: {teams}')

                print(f'Stopping with {len(participants)} participants and {len(teams)} teams left.')
                for team in teams:
                    print(f'Popping team {team.id} with {team.member_count()} participants')
                    print(f'{team.members}')
                    print('\n')

                break  # Break out of the while loop.

            print(f'new round: {rnd}')
            rnd += 1


        pretty_print_teams(full_teams)


if __name__ == '__main__':
    script = Template()
    script.go(sys.argv[1:])
